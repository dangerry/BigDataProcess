#!/usr/bin/python3

from openpyxl import load_workbook
import collections

wb = load_workbook(filename='student.xlsx')
ws = wb['Sheet1']

student = 0
row_id = 1
for row in ws:
    if row_id != 1:
        sum_v = ws.cell(row=row_id, column=3).value * 0.3
        sum_v += ws.cell(row=row_id, column=4).value * 0.35
        sum_v += ws.cell(row=row_id, column=5).value * 0.34
        sum_v += ws.cell(row=row_id, column=6).value
        ws.cell(row=row_id, column=7).value = sum_v
        student += 1
    row_id += 1

id_num = []
total = []
row_id = 1
for row in ws:
    if row_id != 1:
        id_num.append(ws.cell(row=row_id, column=1).value)
        total.append(ws.cell(row=row_id, column=7).value)
    row_id += 1

total_sort = sorted(total, reverse=True)

rank = []
for i in total:
    rank.append(total_sort.index(i)+1)
print(rank)

# same rank count
dict_count = dict(collections.Counter(rank))
print(dict_count)

A = int(student * 0.3); A_plus = int(A * 0.5); A0 = A - A_plus;
B = int(student * 0.7); B_plus = int((B - A) * 0.5); B0 = int(B - A) - B_plus;
C = student; C_plus = int((C - B) * 0.5); C0 = C - B - C_plus
print(A, B, C)
print(A_plus, B_plus, C_plus)
print(A0, B0, C0)

row_id = 2
for ranking in rank:
    if 0 < ranking <= A_plus and A_plus - dict_count[ranking] >= 0:
        ws.cell(row=row_id, column=8).value = 'A+'
        A_plus -= 1
        row_id += 1
    elif A_plus < ranking <= A and A0 - dict_count[ranking] >= 0:
        ws.cell(row=row_id, column=8).value = 'A0'
        A0 -= 1
        row_id += 1
    elif A < ranking <= B - B_plus and B_plus - dict_count[ranking] >= 0:
        ws.cell(row=row_id, column=8).value = 'B+'
        B_plus -= 1
        row_id += 1
    elif B - B_plus < ranking <= B and B0 - dict_count[ranking] >= 0:
        ws.cell(row=row_id, column=8).value = 'B0'
        B0 -= 1
        row_id += 1
    elif B < ranking < C - C_plus and C_plus - dict_count[ranking] >= 0:
        ws.cell(row=row_id, column=8).value = 'C+'
        C_plus -= 1
        row_id += 1
    else:
        ws.cell(row=row_id, column=8).value = 'C0'
        row_id += 1

wb.save("student.xlsx")
